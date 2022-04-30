from openpyxl import load_workbook
import sqlite3
from datetime import datetime
import os


con = sqlite3.connect('engwords.db')
curs = con.cursor()

wb = load_workbook("ready.xlsx")
ws = wb.active
count = 0

def insert_date(eng, rus, trans):
    global count
    try:
        curs.execute(f"""INSERT INTO words (eng_words, rus_words, transcription) VALUES ('{eng}', '{rus}', '{trans}')""")
        con.commit()
    except sqlite3.OperationalError:
        count += 1



def get_words(row):
    return ws[row][0].value, ws[row][2].value


def get_trans(row):
    return ws[row][1].value


def mainloop():
    t = datetime.now()
    for row in range(1, ws.max_row+1):
        insert_date(get_words(row)[0], get_words(row)[1], get_trans(row))
        print("Прогрес:",float(round(row/((ws.max_row+1)/100), 2)), "%")
        os.system('clear')
    print(datetime.now()-t)
    print("Готово")


mainloop()
print("Не добавлено", count, "слов")

